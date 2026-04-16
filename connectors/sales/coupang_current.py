import os
import re
import time
import json
import argparse
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PwTimeoutError

load_dotenv()
KST = timezone(timedelta(hours=9))
MAIN_SELECTOR = "#business-insights-layout__contents__main"
TIMEOUT_5S = 5_000


def must_env(key: str) -> str:
    v = os.getenv(key, "").strip()
    if not v:
        raise RuntimeError(f"{key} 환경변수가 필요합니다. .env를 확인하세요.")
    return v


def kst_today_ymd() -> str:
    return datetime.now(KST).date().strftime("%Y-%m-%d")


def build_sales_url(date_ymd: str) -> str:
    tpl = must_env("COUPANG_SALES_URL_TEMPLATE")
    return tpl.format(date=date_ymd)


def load_brainology_products() -> List[str]:
    raw = os.getenv("BRAINOLOGY_PRODUCT_NAMES", "").strip()
    items = [x.strip() for x in raw.split(",") if x.strip()]
    if not items:
        raise RuntimeError("BRAINOLOGY_PRODUCT_NAMES 환경변수가 필요합니다. 예: 뉴턴젤리,오메가3")
    return items


def product_keywords_map() -> Dict[str, List[str]]:
    result: Dict[str, List[str]] = {}
    for product in load_brainology_products():
        kws = [product]
        alias_env = os.getenv(f"BRAINOLOGY_PRODUCT_ALIASES_{product}", "").strip()
        if alias_env:
            kws.extend([x.strip() for x in alias_env.split(",") if x.strip()])
        result[product] = kws
    return result


def save_debug(page, prefix: str) -> None:
    os.makedirs("debug", exist_ok=True)
    page.screenshot(path=f"debug/{prefix}.png", full_page=True)
    with open(f"debug/{prefix}.html", "w", encoding="utf-8") as f:
        f.write(page.content())


def zero_payload(ymd: str, url: str) -> Dict:
    mapped = {"burdenzero": {"sales": 0, "orders": 0}}
    brand_summary = {"부담제로": {"sales": 0, "qty": 0}}
    for product in load_brainology_products():
        mapped[product] = {"sales": 0, "orders": 0}
        brand_summary[product] = {"sales": 0, "qty": 0}
    return {
        "status": "ok",
        "source": "coupang",
        "date": ymd,
        "target_url": url,
        "excel_path": "",
        "total_sales": 0,
        "total_qty": 0,
        "brand_summary": brand_summary,
        "mapped": mapped,
    }


def wait_quick(page, ms: int = 0) -> None:
    if ms > 0:
        time.sleep(ms / 1000)


def normalize_int(val) -> int:
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(",", "")
    nums = re.findall(r"-?\d+", s)
    return int(nums[0]) if nums else 0


def normalize_text(s: str) -> str:
    return " ".join((s or "").lower().split())


def match_product_key(name: str, keywords: List[str]) -> bool:
    base = normalize_text(name)
    for kw in keywords:
        if normalize_text(kw) in base:
            return True
    return False


def login_coupang(page) -> None:
    login_url = must_env("COUPANG_LOGIN_URL")
    user = must_env("COUPANG_ID")
    pw = must_env("COUPANG_PW")
    page.goto(login_url, wait_until="domcontentloaded")
    try:
        page.wait_for_load_state("networkidle", timeout=15_000)
    except Exception:
        pass
    time.sleep(int(os.getenv("PRE_LOGIN_WAIT_MS", "5000")) / 1000)
    scopes = [page] + list(page.frames)
    submitted = False
    for scope in scopes:
        try:
            id_loc = scope.locator("input[type='email'], input[name='username'], input#username, input[name='id'], input[type='text']")
            pw_loc = scope.locator("input[type='password'], input[name='password'], input#password")
            if id_loc.count() == 0 or pw_loc.count() == 0:
                continue
            id_loc.first.fill(user)
            pw_loc.first.fill(pw)
            btn = scope.locator("button:has-text('로그인'), button:has-text('Login'), button[type='submit'], input[type='submit'], [role='button']:has-text('로그인')")
            if btn.count() > 0:
                btn.first.click()
            else:
                pw_loc.first.press("Enter")
            submitted = True
            break
        except Exception:
            continue
    if not submitted:
        save_debug(page, "coupang_login_form_not_found")
        raise RuntimeError("로그인 폼/버튼을 찾지 못했습니다.")
    time.sleep(3)
    wait_quick(page, int(os.getenv("POST_LOGIN_WAIT_MS", "250")))


def _try_close_overlays_then_wait_main(page) -> bool:
    try:
        try:
            page.get_by_role("button", name="닫기").nth(1).click(timeout=TIMEOUT_5S, force=True)
            wait_quick(page, 100)
        except Exception:
            pass
        try:
            page.get_by_title("closed").nth(17).click(timeout=TIMEOUT_5S, force=True)
            wait_quick(page, 100)
        except Exception:
            pass
        page.wait_for_selector(MAIN_SELECTOR, timeout=TIMEOUT_5S)
        return True
    except Exception:
        return False


def open_sales_url_with_retry(page, url: str, retries: int = 1) -> bool:
    attempts = 0
    while True:
        attempts += 1
        page.goto(url, wait_until="domcontentloaded")
        try:
            page.wait_for_selector(MAIN_SELECTOR, timeout=TIMEOUT_5S)
            return True
        except PwTimeoutError:
            if _try_close_overlays_then_wait_main(page):
                return True
            if attempts > (1 + retries):
                return False
            wait_quick(page, 200)


def download_product_excel_via_dropdown(page, download_dir: str) -> str:
    os.makedirs(download_dir, exist_ok=True)
    page.wait_for_selector(MAIN_SELECTOR, timeout=TIMEOUT_5S)
    page.get_by_text("엑셀 다운로드", exact=True).first.click(timeout=TIMEOUT_5S, force=True)
    menu_item = page.get_by_text("상품별 엑셀 다운로드", exact=True).first
    menu_item.wait_for(state="visible", timeout=TIMEOUT_5S)
    with page.expect_download(timeout=60_000) as d:
        menu_item.click(timeout=TIMEOUT_5S, force=True)
    download = d.value
    path = os.path.join(download_dir, download.suggested_filename)
    download.save_as(path)
    return path


@dataclass
class ProductAgg:
    sales: int = 0
    qty: int = 0


def aggregate_from_excel(path: str) -> Tuple[Dict[str, ProductAgg], int, int]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    agg: Dict[str, ProductAgg] = {}
    total_sales = 0
    total_qty = 0
    c1, o1, p1 = ws["C1"].value, ws["O1"].value, ws["P1"].value
    start_row = 2 if any(isinstance(x, str) for x in (c1, o1, p1)) else 1
    for r in range(start_row, ws.max_row + 1):
        name = ws[f"C{r}"].value
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue
        sales_o = normalize_int(ws[f"O{r}"].value)
        sales_q = normalize_int(ws[f"Q{r}"].value)
        qty_p = normalize_int(ws[f"P{r}"].value)
        qty_r = normalize_int(ws[f"R{r}"].value)
        net_sales = sales_o + sales_q
        net_qty = qty_p + qty_r
        agg.setdefault(name, ProductAgg())
        agg[name].sales += net_sales
        agg[name].qty += net_qty
        total_sales += net_sales
        total_qty += net_qty
    return agg, total_sales, total_qty


def aggregate_by_target(product_agg: Dict[str, ProductAgg]) -> Dict[str, ProductAgg]:
    result: Dict[str, ProductAgg] = {"burdenzero": ProductAgg()}
    kw_map = product_keywords_map()
    for product in kw_map:
        result[product] = ProductAgg()

    for name, agg in product_agg.items():
        if "부담" in name or "부담제로" in name:
            result["burdenzero"].sales += agg.sales
            result["burdenzero"].qty += agg.qty
            continue
        for product, keywords in kw_map.items():
            if match_product_key(name, keywords):
                result[product].sales += agg.sales
                result[product].qty += agg.qty
                break
    return result


def main():
    parser = argparse.ArgumentParser(description="Coupang: product excel download -> net sales/qty -> target summary")
    parser.add_argument("--date", help="집계 날짜 (YYYY-MM-DD). 기본: 오늘(KST)", default=None)
    parser.add_argument("--json", action="store_true", help="러너용: 마지막 줄에 JSON 1줄 출력")
    args = parser.parse_args()
    headless = os.getenv("HEADLESS", "false").lower() == "true"
    ymd = datetime.strptime(args.date, "%Y-%m-%d").date().strftime("%Y-%m-%d") if args.date else kst_today_ymd()
    url = build_sales_url(ymd)

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless)
            context = browser.new_context(accept_downloads=True)
            page = context.new_page()
            try:
                login_coupang(page)
                ok = open_sales_url_with_retry(page, url, retries=1)
                if not ok:
                    payload = zero_payload(ymd, url)
                    print(json.dumps(payload, ensure_ascii=False) if args.json else payload)
                    return
                try:
                    excel_path = download_product_excel_via_dropdown(page, download_dir="downloads")
                except Exception:
                    wait_quick(page, 150)
                    ok2 = open_sales_url_with_retry(page, url, retries=1)
                    if not ok2:
                        payload = zero_payload(ymd, url)
                        print(json.dumps(payload, ensure_ascii=False) if args.json else payload)
                        return
                    excel_path = download_product_excel_via_dropdown(page, download_dir="downloads")

                product_agg, total_sales, total_qty = aggregate_from_excel(excel_path)
                target_agg = aggregate_by_target(product_agg)
                brand_summary = {"부담제로": {"sales": target_agg['burdenzero'].sales, "qty": target_agg['burdenzero'].qty}}
                for product in load_brainology_products():
                    brand_summary[product] = {"sales": target_agg[product].sales, "qty": target_agg[product].qty}
                mapped = {"burdenzero": {"sales": int(target_agg['burdenzero'].sales), "orders": int(target_agg['burdenzero'].qty)}}
                for product in load_brainology_products():
                    mapped[product] = {"sales": int(target_agg[product].sales), "orders": int(target_agg[product].qty)}

                payload = {
                    "status": "ok",
                    "source": "coupang",
                    "date": ymd,
                    "target_url": url,
                    "excel_path": excel_path,
                    "total_sales": int(total_sales),
                    "total_qty": int(total_qty),
                    "brand_summary": brand_summary,
                    "mapped": mapped,
                }
                print(json.dumps(payload, ensure_ascii=False) if args.json else payload)
            except Exception:
                payload = zero_payload(ymd, url)
                print(json.dumps(payload, ensure_ascii=False) if args.json else payload)
            finally:
                try:
                    context.close()
                except Exception:
                    pass
                try:
                    browser.close()
                except Exception:
                    pass
    except Exception:
        payload = zero_payload(ymd, url)
        print(json.dumps(payload, ensure_ascii=False) if args.json else payload)


if __name__ == "__main__":
    main()